"""Read-only MCP server for completed backtest run outputs."""
from __future__ import annotations

import csv
import json
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import load_workbook

from crypto_strategy_lab.run_manifest import (
    RunArtifactError,
    artifact_path,
    load_completed_manifest,
)

LOGGER = logging.getLogger("crypto_strategy_lab.mcp")
READABLE = {".csv", ".tsv", ".xlsx", ".json", ".txt", ".log", ".md", ".parquet"}
MAX_QUERY_ROWS = 5000
MAX_TEXT_BYTES = 1_000_000
FORBIDDEN_SQL = re.compile(
    r"\b(?:INSERT|UPDATE|DELETE|DROP|ALTER|CREATE|COPY|EXPORT|IMPORT|ATTACH|DETACH|"
    r"INSTALL|LOAD|PRAGMA|CALL|SET|RESET|TRUNCATE|VACUUM)\b",
    re.IGNORECASE,
)
ALLOWED_SQL = re.compile(r"^\s*(?:SELECT|WITH|DESCRIBE|SHOW)\b", re.IGNORECASE)
EXTERNAL_SQL = re.compile(
    r"\bread_[a-z0-9_]+\s*\(|"
    r"\b(?:glob|parquet_scan|csv_scan|sqlite_scan|postgres_scan|mysql_scan|delta_scan|iceberg_scan)\s*\(|"
    r"\b(?:FROM|JOIN)\s*['\"]",
    re.IGNORECASE,
)


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, Path)):
        return value.isoformat() if isinstance(value, datetime) else str(value)
    if hasattr(value, "item"):
        return value.item()
    return value


def _validate_sql(sql: str) -> None:
    if ";" in sql:
        raise ValueError("Semicolons and multiple SQL statements are not allowed")
    if not ALLOWED_SQL.match(sql) or FORBIDDEN_SQL.search(sql) or EXTERNAL_SQL.search(sql):
        raise ValueError("Only read-only SELECT, WITH, DESCRIBE, and SHOW queries are allowed")


def _code_identity(manifest: dict[str, Any]) -> tuple[Any, ...]:
    dirty = manifest.get("code_dirty")
    return (
        manifest.get("code_commit"),
        dirty,
        manifest.get("tracked_diff_sha256") if dirty else None,
        tuple(manifest.get("untracked_source_paths", ())) if dirty else (),
    )


class BacktestReports:
    """Validated read-only access to completed runs beneath one output root."""

    def __init__(self, output_root: str | Path):
        root = Path(output_root).expanduser()
        if not root.is_dir():
            raise ValueError(f"Output directory does not exist: {root}")
        self.root = root.resolve(strict=True)

    def _inside(self, path: Path) -> bool:
        return path == self.root or self.root in path.parents

    def _resolve(self, relative: str | Path, *, directory: bool | None = None) -> Path:
        raw = Path(relative)
        if raw.is_absolute():
            raise ValueError("Absolute paths are not allowed")
        if not raw.parts or any(part == ".." for part in raw.parts):
            raise ValueError("Path traversal is not allowed")
        try:
            resolved = (self.root / raw).resolve(strict=True)
        except (FileNotFoundError, RuntimeError) as exc:
            raise ValueError(f"Path does not exist: {relative}") from exc
        if not self._inside(resolved):
            raise ValueError("Resolved path is outside the allowed output root")
        if directory is True and not resolved.is_dir():
            raise ValueError("Run is not a directory")
        if directory is False and not resolved.is_file():
            raise ValueError("Report is not a file")
        return resolved

    def resolve_run(self, run: str) -> Path:
        path = self._resolve(run, directory=True)
        if path.parent != self.root:
            raise ValueError("Run must be a direct child of the output root")
        return path

    @staticmethod
    def _load_json(path: Path) -> dict[str, Any]:
        try:
            value = json.loads(path.read_text(encoding="utf-8-sig"))
            return value if isinstance(value, dict) else {}
        except (OSError, json.JSONDecodeError):
            return {}

    @staticmethod
    def _metadata(run: Path, manifest: dict[str, Any]) -> dict[str, Any]:
        request = manifest["request"]
        hashes = manifest["hashes"]
        return {
            "folder_name": run.name,
            "run_id": manifest["run_id"],
            "run_started_at": manifest["run_started_at"],
            "run_completed_at": manifest["run_completed_at"],
            "symbol": request["symbol"],
            "start": request["start"],
            "end": request["end"],
            "strategy_timeframe": request["requested_strategy_interval"],
            "execution_timeframe": request["effective_intrabar_interval"],
            "trade_count": manifest["execution_result"]["completed_trade_rows"],
            "strategy_hash": hashes["strategy_hash"],
            "execution_hash": hashes["execution_hash"],
            "code_commit": manifest.get("code_commit"),
            "reproducibility_status": manifest.get("reproducibility_status"),
        }

    def list_runs(self, limit: int = 50) -> list[dict[str, Any]]:
        if not 1 <= limit <= 1000:
            raise ValueError("limit must be between 1 and 1000")
        runs: list[dict[str, Any]] = []
        for path in self.root.iterdir():
            if not path.is_dir() or path.is_symlink():
                continue
            try:
                manifest = load_completed_manifest(path)
            except RunArtifactError:
                continue
            runs.append(self._metadata(path, manifest))
        runs.sort(key=lambda item: item["run_started_at"], reverse=True)
        return runs[:limit]

    def latest_run(self) -> dict[str, Any]:
        runs = self.list_runs(1)
        if not runs:
            raise ValueError("No completed backtest runs were found")
        run = self.resolve_run(runs[0]["folder_name"])
        manifest = load_completed_manifest(run)
        result = {
            **runs[0],
            "path": run.name,
            "artifact_availability": {name: True for name in manifest["artifacts"]},
        }
        result["summary"] = self._load_json(artifact_path(run, manifest, "summary"))
        return result

    def get_run_manifest(self, run: str) -> dict[str, Any]:
        return load_completed_manifest(self.resolve_run(run))

    @staticmethod
    def _artifact_by_path(manifest: dict[str, Any]) -> dict[str, tuple[str, dict[str, Any]]]:
        return {
            str(item["path"]): (name, item)
            for name, item in manifest.get("artifacts", {}).items()
        }

    @staticmethod
    def _reject_symlink_components(folder: Path, relative: Path) -> None:
        current = folder
        for part in relative.parts:
            current = current / part
            if current.is_symlink():
                raise ValueError("Symlinked run files are not allowed")

    def _run_file_path(self, run: str, filename: str) -> Path:
        folder = self.resolve_run(run)
        manifest = load_completed_manifest(folder)
        relative = Path(filename)
        if relative.is_absolute():
            raise ValueError("Absolute paths are not allowed")
        if not relative.parts or any(part == ".." for part in relative.parts):
            raise ValueError("Path traversal is not allowed")
        self._reject_symlink_components(folder, relative)
        candidate = folder / relative
        try:
            resolved = candidate.resolve(strict=True)
        except (FileNotFoundError, RuntimeError) as exc:
            raise ValueError(f"Path does not exist: {filename}") from exc
        if resolved == folder or folder not in resolved.parents or not resolved.is_file():
            raise ValueError("Run file must be a regular file inside the selected run")

        registered = self._artifact_by_path(manifest).get(relative.as_posix())
        if registered is not None:
            artifact_name, _ = registered
            return artifact_path(folder, manifest, artifact_name)
        return resolved

    def list_run_files(self, run: str) -> list[dict[str, Any]]:
        folder = self.resolve_run(run)
        manifest = load_completed_manifest(folder)
        artifact_lookup = self._artifact_by_path(manifest)
        result: list[dict[str, Any]] = []
        for path in folder.rglob("*"):
            if path.is_symlink() or not path.is_file():
                continue
            relative = path.relative_to(folder).as_posix()
            registered = artifact_lookup.get(relative)
            artifact_name, item = registered if registered else (None, None)
            suffix = path.suffix.lower()
            result.append(
                {
                    "artifact": artifact_name,
                    "filename": relative,
                    "extension": suffix,
                    "size": path.stat().st_size,
                    "sha256": item.get("sha256") if item else None,
                    "registered_artifact": registered is not None,
                    "readable": suffix in READABLE,
                }
            )
        return sorted(result, key=lambda item: item["filename"])

    @staticmethod
    def _table_payload(columns, rows, total: int, limit: int, **extra) -> dict[str, Any]:
        return {
            "type": "table",
            **extra,
            "columns": list(columns),
            "rows": rows,
            "row_count": total,
            "truncated": total > limit,
        }

    def _preview_parquet(self, path: Path, limit: int) -> dict[str, Any]:
        with duckdb.connect(":memory:") as connection:
            escaped = str(path).replace("'", "''")
            connection.execute(f"CREATE VIEW data AS SELECT * FROM read_parquet('{escaped}')")
            total = int(connection.execute("SELECT count(*) FROM data").fetchone()[0])
            cursor = connection.execute("SELECT * FROM data LIMIT ?", [limit])
            columns = [item[0] for item in (cursor.description or [])]
            rows = [[_json_value(value) for value in row] for row in cursor.fetchall()]
        return self._table_payload(columns, rows, total, limit, format="parquet")

    def read_report(
        self,
        run: str,
        filename: str,
        sheet: str | None = None,
        limit: int = 200,
    ) -> dict[str, Any]:
        """Read any supported file inside a completed run; kept for MCP compatibility."""
        if not 1 <= limit <= MAX_QUERY_ROWS:
            raise ValueError(f"limit must be between 1 and {MAX_QUERY_ROWS}")
        path = self._run_file_path(run, filename)
        suffix = path.suffix.lower()
        if suffix not in READABLE:
            raise ValueError(
                "Unsupported run file type; readable types are: "
                + ", ".join(sorted(READABLE))
            )
        if suffix == ".json":
            return {"type": "json", "data": json.loads(path.read_text(encoding="utf-8-sig"))}
        if suffix in {".txt", ".log", ".md"}:
            with path.open("r", encoding="utf-8-sig", errors="replace") as handle:
                text = handle.read(MAX_TEXT_BYTES + 1)
            return {
                "type": "text",
                "text": text[:MAX_TEXT_BYTES],
                "truncated": len(text) > MAX_TEXT_BYTES,
            }
        if suffix in {".csv", ".tsv"}:
            delimiter = "\t" if suffix == ".tsv" else ","
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle, delimiter=delimiter)
                columns = next(reader, [])
                rows, total = [], 0
                for row in reader:
                    total += 1
                    if len(rows) < limit:
                        rows.append(row)
            return self._table_payload(columns, rows, total, limit)
        if suffix == ".parquet":
            return self._preview_parquet(path, limit)

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet is not None and sheet not in workbook.sheetnames:
                raise ValueError(f"Unknown sheet; available sheets: {workbook.sheetnames}")
            ws = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
            iterator = ws.iter_rows(values_only=True)
            columns = [str(value) if value is not None else "" for value in next(iterator, ())]
            rows, total = [], 0
            for row in iterator:
                total += 1
                if len(rows) < limit:
                    rows.append([_json_value(value) for value in row])
            return self._table_payload(
                columns,
                rows,
                total,
                limit,
                sheet=ws.title,
                available_sheets=workbook.sheetnames,
            )
        finally:
            workbook.close()

    def read_run_file(
        self, run: str, filename: str, sheet: str | None = None, limit: int = 200
    ) -> dict[str, Any]:
        return self.read_report(run, filename, sheet, limit)

    @staticmethod
    def _query_parquet_path(path: Path, relation: str, sql: str) -> dict[str, Any]:
        _validate_sql(sql)
        with duckdb.connect(":memory:") as connection:
            escaped = str(path).replace("'", "''")
            connection.execute(
                f"CREATE VIEW {relation} AS SELECT * FROM read_parquet('{escaped}')"
            )
            cursor = connection.execute(sql)
            columns = [item[0] for item in (cursor.description or [])]
            rows = cursor.fetchmany(MAX_QUERY_ROWS + 1)
        truncated = len(rows) > MAX_QUERY_ROWS
        rows = rows[:MAX_QUERY_ROWS]
        return {
            "columns": columns,
            "rows": [[_json_value(value) for value in row] for row in rows],
            "row_count": len(rows),
            "truncated": truncated,
        }

    def _query_artifact_parquet(
        self, run: str, artifact: str, relation: str, sql: str
    ) -> dict[str, Any]:
        folder = self.resolve_run(run)
        manifest = load_completed_manifest(folder)
        path = artifact_path(folder, manifest, artifact)
        return self._query_parquet_path(path, relation, sql)

    def query_trades(self, run: str, sql: str) -> dict[str, Any]:
        return self._query_artifact_parquet(run, "trades", "trades", sql)

    def query_signals(self, run: str, sql: str) -> dict[str, Any]:
        return self._query_artifact_parquet(run, "signals", "signals", sql)

    def query_feature_context(self, run: str, sql: str) -> dict[str, Any]:
        return self._query_artifact_parquet(run, "feature_context", "feature_context", sql)

    def query_parquet(self, run: str, filename: str, sql: str) -> dict[str, Any]:
        path = self._run_file_path(run, filename)
        if path.suffix.lower() != ".parquet":
            raise ValueError("query_parquet requires a .parquet run file")
        return self._query_parquet_path(path, "data", sql)

    def research_aggregate(self, run: str, spec: dict[str, Any]) -> dict[str, Any]:
        from crypto_strategy_lab.feature_research import ResearchQueryService

        with ResearchQueryService(self.resolve_run(run)) as service:
            frame = service.query(spec)
        total = len(frame)
        truncated = total > MAX_QUERY_ROWS
        frame = frame.head(MAX_QUERY_ROWS)
        return {
            "columns": list(frame.columns),
            "rows": [
                {key: _json_value(value) for key, value in row.items()}
                for row in frame.to_dict(orient="records")
            ],
            "row_count": len(frame),
            "truncated": truncated,
        }

    def compare_runs(self, runs: list[str]) -> list[dict[str, Any]]:
        if not 2 <= len(runs) <= 10 or len(set(runs)) != len(runs):
            raise ValueError("Provide 2-10 distinct run folder names")
        result: list[dict[str, Any]] = []
        manifests: list[dict[str, Any]] = []
        for name in runs:
            folder = self.resolve_run(name)
            manifest = load_completed_manifest(folder)
            manifests.append(manifest)
            summary = self._load_json(artifact_path(folder, manifest, "summary"))
            initial_equity = float(
                manifest.get("config", {}).get("execution", {}).get("initial_equity", 0.0) or 0.0
            )
            net_pnl = summary.get("net_pnl")
            if net_pnl is None and summary.get("ending_equity") is not None:
                net_pnl = float(summary["ending_equity"]) - initial_equity
            result.append(
                {
                    "run": name,
                    "symbol": manifest["request"]["symbol"],
                    **manifest["hashes"],
                    "code_commit": manifest.get("code_commit"),
                    "source_snapshot": manifest["catalog"]["catalog_snapshot_digest"],
                    "total_trades": summary.get("total_trades"),
                    "win_rate": summary.get("win_rate"),
                    "net_r": summary.get("total_net_r"),
                    "avg_r": summary.get("average_net_r"),
                    "net_pnl": net_pnl,
                }
            )
        baseline = manifests[0]
        baseline_code = _code_identity(baseline)
        for row, manifest in zip(result, manifests):
            row["provenance_vs_first"] = {
                "same_code": _code_identity(manifest) == baseline_code,
                "same_sources": manifest["catalog"]["catalog_snapshot_digest"]
                == baseline["catalog"]["catalog_snapshot_digest"],
                "same_features": manifest["hashes"]["feature_config_hash"]
                == baseline["hashes"]["feature_config_hash"]
                and manifest["features"] == baseline["features"],
                "same_strategy": manifest["hashes"]["strategy_hash"]
                == baseline["hashes"]["strategy_hash"],
                "same_execution": manifest["hashes"]["execution_hash"]
                == baseline["hashes"]["execution_hash"],
            }
        return result


def resolve_output_root() -> Path:
    configured = os.environ.get("CRYPTO_STRATEGY_LAB_OUTPUT_DIR")
    candidate = (
        Path(configured).expanduser()
        if configured
        else Path(__file__).resolve().parents[1] / "output"
    )
    if not candidate.is_dir():
        source = "CRYPTO_STRATEGY_LAB_OUTPUT_DIR" if configured else "the project's default output directory"
        raise RuntimeError(
            f"Could not resolve {source}: {candidate}. Create it or set CRYPTO_STRATEGY_LAB_OUTPUT_DIR."
        )
    return candidate.resolve(strict=True)


def create_server(reports: BacktestReports):
    from mcp.server import MCPServer

    server = MCPServer("Crypto Strategy Lab Reports")

    @server.tool()
    def list_runs(limit: int = 50) -> list[dict[str, Any]]:
        LOGGER.info("MCP tool called: list_runs")
        return reports.list_runs(limit)

    @server.tool()
    def latest_run() -> dict[str, Any]:
        LOGGER.info("MCP tool called: latest_run")
        return reports.latest_run()

    @server.tool()
    def get_run_manifest(run: str) -> dict[str, Any]:
        LOGGER.info("MCP tool called: get_run_manifest run=%s", run)
        return reports.get_run_manifest(run)

    @server.tool()
    def list_run_files(run: str) -> list[dict[str, Any]]:
        LOGGER.info("MCP tool called: list_run_files run=%s", run)
        return reports.list_run_files(run)

    @server.tool()
    def read_report(
        run: str, filename: str, sheet: str | None = None, limit: int = 200
    ) -> dict[str, Any]:
        LOGGER.info("MCP tool called: read_report run=%s filename=%s", run, filename)
        return reports.read_report(run, filename, sheet, limit)

    @server.tool()
    def read_run_file(
        run: str, filename: str, sheet: str | None = None, limit: int = 200
    ) -> dict[str, Any]:
        LOGGER.info("MCP tool called: read_run_file run=%s filename=%s", run, filename)
        return reports.read_run_file(run, filename, sheet, limit)

    @server.tool()
    def query_trades(run: str, sql: str) -> dict[str, Any]:
        LOGGER.info("MCP tool called: query_trades run=%s", run)
        return reports.query_trades(run, sql)

    @server.tool()
    def query_signals(run: str, sql: str) -> dict[str, Any]:
        LOGGER.info("MCP tool called: query_signals run=%s", run)
        return reports.query_signals(run, sql)

    @server.tool()
    def query_feature_context(run: str, sql: str) -> dict[str, Any]:
        LOGGER.info("MCP tool called: query_feature_context run=%s", run)
        return reports.query_feature_context(run, sql)

    @server.tool()
    def query_parquet(run: str, filename: str, sql: str) -> dict[str, Any]:
        LOGGER.info("MCP tool called: query_parquet run=%s filename=%s", run, filename)
        return reports.query_parquet(run, filename, sql)

    @server.tool()
    def research_aggregate(run: str, spec: dict[str, Any]) -> dict[str, Any]:
        LOGGER.info("MCP tool called: research_aggregate run=%s", run)
        return reports.research_aggregate(run, spec)

    @server.tool()
    def compare_runs(runs: list[str]) -> list[dict[str, Any]]:
        LOGGER.info("MCP tool called: compare_runs count=%d", len(runs))
        return reports.compare_runs(runs)

    return server


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    root = resolve_output_root()
    host = "127.0.0.1"
    try:
        port = int(os.environ.get("CRYPTO_STRATEGY_LAB_MCP_PORT", "8765"))
        if not 1 <= port <= 65535:
            raise ValueError
    except ValueError as exc:
        raise SystemExit("CRYPTO_STRATEGY_LAB_MCP_PORT must be an integer from 1 to 65535") from exc
    LOGGER.info("MCP server starting")
    LOGGER.info("Allowed output root: %s", root)
    LOGGER.info("Host: %s", host)
    LOGGER.info("Port: %s", port)
    LOGGER.info(
        "Available tools: list_runs, latest_run, get_run_manifest, list_run_files, "
        "read_report, read_run_file, query_trades, query_signals, query_feature_context, "
        "query_parquet, research_aggregate, compare_runs"
    )
    create_server(BacktestReports(root)).run(
        transport="streamable-http",
        host=host,
        port=port,
        streamable_http_path="/mcp",
    )


if __name__ == "__main__":
    main()
