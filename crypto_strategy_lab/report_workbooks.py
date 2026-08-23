"""Excel workbook writers for human-facing backtest reports.

Calculation remains in the existing statistics and analysis modules.  This
module is deliberately limited to arranging already-built tables and applying
consistent, lightweight spreadsheet formatting.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Mapping

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _excel_safe_value(value: Any) -> Any:
    """Return an Excel-compatible scalar while preserving the displayed UTC time."""
    if isinstance(value, pd.Timestamp) and value.tzinfo is not None:
        return value.tz_convert("UTC").tz_localize(None)
    return value


def _excel_safe_frame(frame: pd.DataFrame | None) -> pd.DataFrame:
    """Strip timezone metadata from datetime cells before handing data to openpyxl.

    Excel has no timezone-aware datetime type.  We keep the same UTC wall-clock
    value and remove only the timezone metadata. This is an export-only copy and
    never changes the calculations or source DataFrame.
    """
    if frame is None:
        return pd.DataFrame()
    result = frame.copy()
    for column in result.columns:
        series = result[column]
        if isinstance(series.dtype, pd.DatetimeTZDtype):
            result[column] = series.dt.tz_convert("UTC").dt.tz_localize(None)
        elif series.dtype == object:
            result[column] = series.map(_excel_safe_value)
    return result


def _format_table_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    if sheet.max_row and sheet.max_column:
        sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for column in range(1, sheet.max_column + 1):
        values = [str(sheet.cell(row, column).value or "") for row in range(1, min(sheet.max_row, 200) + 1)]
        sheet.column_dimensions[get_column_letter(column)].width = min(42, max(10, max(map(len, values), default=0) + 2))
        heading = str(sheet.cell(1, column).value or "").lower()
        number_format = None
        if "rate" in heading or "percentage" in heading or heading.endswith(" %"):
            number_format = "0.00%"
        elif any(token in heading for token in ("pnl", "equity", "profit", "drawdown")):
            number_format = '#,##0.00;[Red]-#,##0.00'
        elif heading in {"avg_r", "average_r", "total_r", "net_r"} or heading.endswith(" r"):
            number_format = '0.0000"R";[Red]-0.0000"R"'
        if number_format:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, column).number_format = number_format


def _write_tables(path: Path, tables: Mapping[str, pd.DataFrame]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in tables.items():
            _excel_safe_frame(frame).to_excel(writer, sheet_name=sheet_name, index=False)
        for sheet in writer.book.worksheets:
            _format_table_sheet(sheet)
    return path


def build_indicator_workbook(tables: Mapping[str, pd.DataFrame], run_dir: Path) -> Path:
    """Write extensible indicator tables; callers may add another named sheet."""
    return _write_tables(run_dir / "indicator_analysis.xlsx", tables)


def build_periodic_breakdown(trades: pd.DataFrame, freq: str) -> pd.DataFrame:
    """Summarize the current single-side trade contract by exit period."""
    columns = ["period", "pair_count", "net_pnl", "net_r"]
    if trades.empty:
        return pd.DataFrame(columns=columns)
    required = {"exit_time", "pair_net_pnl", "pair_net_r"}
    missing = sorted(required - set(trades.columns))
    if missing:
        raise ValueError(f"Periodic report is missing current trade columns: {missing}")

    exits = pd.to_datetime(trades["exit_time"], errors="coerce", utc=True)
    if exits.isna().all():
        raise ValueError("Periodic results require at least one valid trade exit timestamp.")
    frame = pd.DataFrame(
        {
            "exit_time": exits,
            "pair_net_pnl": pd.to_numeric(trades["pair_net_pnl"], errors="coerce"),
            "pair_net_r": pd.to_numeric(trades["pair_net_r"], errors="coerce"),
        }
    )
    frame = frame.loc[frame["exit_time"].notna()].set_index("exit_time")
    candidates = [freq]
    fallback = {"ME": "M", "YE": "Y", "M": "ME", "Y": "YE"}.get(freq)
    if fallback is not None:
        candidates.append(fallback)
    last_error: Exception | None = None
    for candidate in candidates:
        try:
            periodic = frame.resample(candidate).agg(
                pair_count=("pair_net_pnl", "size"),
                net_pnl=("pair_net_pnl", "sum"),
                net_r=("pair_net_r", "sum"),
            )
            return periodic.reset_index().rename(columns={"exit_time": "period"})
        except ValueError as exc:
            last_error = exc
    raise last_error if last_error is not None else ValueError(f"Unsupported resample frequency: {freq}")


def build_performance_breakdowns(trades: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build regime and regime-by-side tables from the modern trade schema."""
    if trades.empty:
        return pd.DataFrame(), pd.DataFrame()

    required = {"market_regime", "side", "pair_net_pnl", "pair_net_r"}
    missing = sorted(required - set(trades.columns))
    if missing:
        raise ValueError(f"Performance report is missing current trade columns: {missing}")

    def aggregate(frame: pd.DataFrame, groups: list[str]) -> pd.DataFrame:
        rows = []
        for keys, group in frame.groupby(groups, dropna=False, observed=True):
            keys = keys if isinstance(keys, tuple) else (keys,)
            pnl_values = pd.to_numeric(group["pair_net_pnl"], errors="coerce")
            r_values = pd.to_numeric(group["pair_net_r"], errors="coerce")
            row = dict(zip(groups, keys))
            row.update(
                {
                    "trades": len(group),
                    "wins": int((pnl_values > 0).sum()),
                    "losses": int((pnl_values < 0).sum()),
                    "win_rate": float((pnl_values > 0).mean()),
                    "net_pnl": float(pnl_values.sum()),
                    "average_r": float(r_values.mean()),
                    "total_r": float(r_values.sum()),
                }
            )
            rows.append(row)
        return pd.DataFrame(rows)

    market = aggregate(trades, ["market_regime"])
    directional = trades.loc[trades["side"].notna()].copy()
    if directional.empty:
        return market, pd.DataFrame()
    directional["direction"] = directional["side"].astype("string").str.upper()
    return market, aggregate(directional, ["market_regime", "direction"])


def _dashboard_metrics(summary: Mapping[str, Any], config: Any, run_dir: Path) -> list[tuple[str, Any]]:
    candidates = [
        ("Run Name", getattr(config, "run_name", None) or run_dir.name),
        ("Symbol", getattr(config, "input_csv", None) and Path(config.input_csv).stem),
        ("Strategy Timeframe", getattr(config, "strategy_timeframe_minutes", None)),
        ("Intrabar Timeframe", getattr(config, "intrabar_timeframe_minutes", None) if getattr(config, "use_intrabar_data", False) else None),
        ("Trading Start", summary.get("trading_start")), ("Trading End", summary.get("trading_end")),
        ("Initial Equity", getattr(config, "initial_equity", None)), ("Final Equity", summary.get("ending_equity")),
        ("Net PnL", (summary.get("ending_equity") - config.initial_equity) if summary.get("ending_equity") is not None else None),
        ("Return %", summary.get("total_return_percentage")), ("Trades", summary.get("total_trades", summary.get("total_pairs"))),
        ("Wins", summary.get("wins")), ("Losses", summary.get("losses")), ("Win Rate", summary.get("win_rate")),
        ("Average R", summary.get("average_net_r")), ("Total R", summary.get("total_net_r")),
        ("Profit Factor", summary.get("profit_factor")), ("Max Drawdown", summary.get("maximum_drawdown")),
    ]
    return [(label, value) for label, value in candidates if value is not None]


def build_backtest_workbook(summary: Mapping[str, Any], config: Any, run_dir: Path,
                            monthly: pd.DataFrame, yearly: pd.DataFrame,
                            market_regime: pd.DataFrame | None = None,
                            direction_regime: pd.DataFrame | None = None) -> Path:
    path = run_dir / "backtest_report.xlsx"
    tables = {"Monthly": monthly, "Yearly": yearly,
              "Market Regime": market_regime if market_regime is not None else pd.DataFrame(),
              "Direction - Regime": direction_regime if direction_regime is not None else pd.DataFrame()}
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        dashboard_frame = pd.DataFrame(_dashboard_metrics(summary, config, run_dir), columns=["Metric", "Value"])
        _excel_safe_frame(dashboard_frame).to_excel(writer, sheet_name="Dashboard", index=False)
        for name, frame in tables.items():
            _excel_safe_frame(frame).to_excel(writer, sheet_name=name, index=False)
        for sheet in writer.book.worksheets:
            _format_table_sheet(sheet)
        dashboard = writer.book["Dashboard"]
        dashboard.column_dimensions["A"].width = 28
        dashboard.column_dimensions["B"].width = 28
    return path


def build_support_resistance_workbook(tables: Mapping[str, pd.DataFrame], run_dir: Path) -> Path:
    order = ["Overview", "Location", "Distance", "Regime", "Event Context", "Hold Analysis", "Rejection Strength", "Test Count"]
    normalized = {name: tables.get(name, pd.DataFrame()) for name in order}
    return _write_tables(run_dir / "support_resistance_analysis.xlsx", normalized)
