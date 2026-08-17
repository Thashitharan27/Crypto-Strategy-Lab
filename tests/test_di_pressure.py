import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from crypto_strategy_lab.config import BacktestConfig
from crypto_strategy_lab.engine import BacktestEngine
from crypto_strategy_lab.report_workbooks import build_indicator_workbook
from crypto_strategy_lab.statistics import di_pressure_analysis


def candles(n=20):
    return pd.DataFrame({"timestamp":pd.date_range("2024-01-01",periods=n,freq="15min",tz="UTC"),"open":np.arange(n)+100.,"high":np.arange(n)+102.,"low":np.arange(n)+99.,"close":np.arange(n)+101.,"volume":1.})

def engine(plus,minus,**kwargs):
    e=BacktestEngine(candles(len(plus)),BacktestConfig(adx_period=2,di_pressure_lookback=3,**kwargs))
    e.plus_di_values[:]=plus;e.minus_di_values[:]=minus;e.di_spread[:]=np.abs(np.array(plus)-np.array(minus));return e

@pytest.mark.parametrize("plus,minus,direction,state,dc,oc",[
([30]*3+[36],[20]*3+[15],"LONG","EXPANDING",6,-5),
([36]*3+[30],[15]*3+[20],"LONG","CONTRACTING",-6,5),
([22]*3+[16],[28]*3+[36],"SHORT","EXPANDING",8,-6),
([16]*3+[22],[36]*3+[28],"SHORT","CONTRACTING",-8,6),
([20]*3+[25],[10]*3+[15],"LONG","MIXED",5,5),
([20]*3+[15],[10]*3+[5],"LONG","MIXED",-5,-5),
([20]*4,[10]*4,"LONG","MIXED",0,0),
])
def test_pressure_classification(plus,minus,direction,state,dc,oc):
    value=engine(plus,minus)._di_pressure_snapshot(3,direction)
    assert value["di_pressure_state"]==state
    assert value["directional_di_change"]==dc and value["opposing_di_change"]==oc

def test_di_direction_selection_and_insufficient_history():
    e=engine([1,1,30,10],[1,1,10,30])
    assert e._selected_direction(2)=="LONG" and e._selected_direction(3)=="SHORT"
    assert e._di_pressure_snapshot(2,"LONG")["di_pressure_state"]=="UNKNOWN"

def test_pressure_uses_no_future_values():
    e=engine([10,11,12,20,999],[20,19,18,10,0])
    before=e._di_pressure_snapshot(3,"LONG")
    e.plus_di_values[4]=-999;e.minus_di_values[4]=999
    assert e._di_pressure_snapshot(3,"LONG")==before

def test_pressure_toggle_changes_only_telemetry():
    on=engine([10,11,12,20],[20,19,18,10]); off=engine([10,11,12,20],[20,19,18,10],enable_di_pressure_analysis=False)
    assert on._selected_direction(3)==off._selected_direction(3)=="LONG"
    assert on._di_pressure_snapshot(3,"LONG")["di_pressure_state"]=="EXPANDING"
    assert off._di_pressure_snapshot(3,"LONG")["di_pressure_state"]=="UNKNOWN"

def test_report_sheet_and_totals(tmp_path):
    trades=pd.DataFrame({"side":["LONG","SHORT"],"market_regime":["BULL","BEAR"],"di_pressure_state":["EXPANDING","MIXED"],"di_spread_change":[6,-6],"pair_net_pnl":[10,-5],"pair_net_r":[1,-.5]})
    table=di_pressure_analysis(trades)
    build_indicator_workbook({"DI Pressure":table},tmp_path)
    assert load_workbook(tmp_path/"indicator_analysis.xlsx",read_only=True).sheetnames==["DI Pressure"]
    assert table.loc[table.Section=="Direction + Pressure","Trades"].sum()==len(trades)
