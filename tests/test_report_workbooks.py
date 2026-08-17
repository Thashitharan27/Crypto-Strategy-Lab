import pandas as pd
from openpyxl import load_workbook

from crypto_strategy_lab.config import BacktestConfig
from crypto_strategy_lab.gui.main_window import report_button_states
from crypto_strategy_lab.report_workbooks import build_backtest_workbook, build_indicator_workbook
from crypto_strategy_lab.support_resistance_analysis import generate_sr_analysis_reports


def test_backtest_and_indicator_workbooks_have_expected_sheets(tmp_path):
    config = BacktestConfig(initial_equity=1000)
    build_backtest_workbook({"ending_equity": 1010, "total_pairs": 1}, config, tmp_path,
                            pd.DataFrame({"period": ["Jan"]}), pd.DataFrame({"period": ["2026"]}))
    build_indicator_workbook({"ADX": pd.DataFrame(), "BB Width": pd.DataFrame(), "DI Spread": pd.DataFrame()}, tmp_path)
    assert load_workbook(tmp_path / "backtest_report.xlsx", read_only=True).sheetnames == ["Dashboard", "Monthly", "Yearly", "Market Regime", "Direction - Regime"]
    assert load_workbook(tmp_path / "indicator_analysis.xlsx", read_only=True).sheetnames == ["ADX", "BB Width", "DI Spread"]


def test_empty_optional_sr_analysis_still_creates_workbook(tmp_path):
    generate_sr_analysis_reports(pd.DataFrame(), tmp_path)
    assert (tmp_path / "support_resistance_analysis.xlsx").exists()


def test_report_button_states_follow_file_existence(tmp_path):
    (tmp_path / "trade_list.csv").write_text("trade\n")
    (tmp_path / "charts").mkdir()
    states = report_button_states(tmp_path)
    assert states["trades"] and states["charts"] and states["output"]
    assert not states["backtest"] and not states["indicators"] and not states["sr"]
