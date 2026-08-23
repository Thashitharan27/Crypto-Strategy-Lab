from datetime import datetime, timezone
from types import SimpleNamespace

import pandas as pd
from openpyxl import load_workbook

from crypto_strategy_lab.data.query import DataRequest
from crypto_strategy_lab.data.quality import (
    DataQualityStatus,
    VALIDATION_CONTRACT_VERSION,
    validate_dataset,
)
from crypto_strategy_lab.data.schemas import DatasetKind
from crypto_strategy_lab.research_reporting import _build_human_workbook
from crypto_strategy_lab.report_workbooks import build_performance_breakdowns


def _modern_trades() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "exit_time": pd.to_datetime(
                [
                    "2025-01-15T04:00:00Z",
                    "2025-02-20T08:00:00Z",
                    "2026-01-10T12:00:00Z",
                ],
                utc=True,
            ),
            "pair_net_pnl": [10.0, -5.0, 7.0],
            "pair_net_r": [1.0, -0.5, 0.7],
            "market_regime": ["BULL", "BEAR", "BULL"],
            "side": ["LONG", "SHORT", "SHORT"],
        }
    )


def test_canonical_human_workbook_populates_all_breakdown_sheets(tmp_path):
    trades = _modern_trades()
    config = SimpleNamespace(
        run_name="report-test",
        input_csv="BTCUSDT.csv",
        strategy_timeframe_minutes=240,
        intrabar_timeframe_minutes=1,
        use_intrabar_data=True,
        initial_equity=1000.0,
    )
    path = _build_human_workbook(
        {"ending_equity": 1012.0, "total_trades": len(trades)},
        config,
        tmp_path,
        trades,
    )

    workbook = load_workbook(path, read_only=True)
    for sheet_name in ("Monthly", "Yearly", "Market Regime", "Direction - Regime"):
        assert workbook[sheet_name].max_row > 1, f"{sheet_name} should contain data rows"


def test_regime_direction_breakdown_uses_modern_side_column():
    market, direction = build_performance_breakdowns(_modern_trades())

    assert set(market["market_regime"]) == {"BULL", "BEAR"}
    assert set(direction["direction"]) == {"LONG", "SHORT"}
    bull_long = direction.loc[
        direction["market_regime"].eq("BULL") & direction["direction"].eq("LONG")
    ].iloc[0]
    assert bull_long["trades"] == 1
    assert bull_long["wins"] == 1
    assert bull_long["total_r"] == 1.0


def _premium_frame() -> tuple[DataRequest, pd.DataFrame]:
    start = datetime(2026, 1, 1, tzinfo=timezone.utc)
    end = datetime(2026, 1, 1, 4, tzinfo=timezone.utc)
    request = DataRequest(
        symbol="BTCUSDT",
        start=start,
        end=end,
        strategy_interval="1h",
        datasets=(DatasetKind.PREMIUM_INDEX_KLINES,),
    )
    starts = pd.date_range(start, end, freq="1h", inclusive="left")
    frame = pd.DataFrame(
        {
            "period_start": starts,
            "period_end": starts + pd.Timedelta("1h"),
            "available_at": starts + pd.Timedelta("1h"),
            "open": [-0.00040, -0.00030, -0.00010, -0.00020],
            "high": [-0.00020, 0.00000, 0.00010, -0.00005],
            "low": [-0.00060, -0.00050, -0.00030, -0.00040],
            "close": [-0.00045, -0.00010, 0.00000, -0.00025],
            "volume": [0.0, 0.0, 0.0, 0.0],
            "symbol": request.symbol,
            "exchange": request.exchange,
            "market": request.market.value,
            "dataset": DatasetKind.PREMIUM_INDEX_KLINES.value,
            "interval": "1h",
        }
    )
    frame.attrs["canonical_source_identity"] = "premium-source"
    return request, frame


def test_premium_index_allows_signed_and_zero_ohlc_values():
    request, frame = _premium_frame()
    report = validate_dataset(
        frame,
        request,
        DatasetKind.PREMIUM_INDEX_KLINES,
        interval="1h",
    )

    assert VALIDATION_CONTRACT_VERSION == "4"
    assert report.status is DataQualityStatus.OK
    assert not any(issue.code == "INVALID_DOMAIN_VALUE" for issue in report.issues)


def test_premium_index_still_reports_real_timeline_gaps():
    request, frame = _premium_frame()
    report = validate_dataset(
        frame.drop(index=[2]).reset_index(drop=True),
        request,
        DatasetKind.PREMIUM_INDEX_KLINES,
        interval="1h",
        required=False,
    )

    assert report.status is DataQualityStatus.WARN
    assert any(issue.code == "MISSING_INTERNAL_INTERVAL" for issue in report.issues)
