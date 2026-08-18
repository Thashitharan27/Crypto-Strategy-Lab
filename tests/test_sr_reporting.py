"""Tests for entry-time S/R event classification and the S/R analysis report files."""
from types import SimpleNamespace

import pandas as pd
from openpyxl import load_workbook

from crypto_strategy_lab.config import BacktestConfig, EntryMode, RiskMode
from crypto_strategy_lab.engine import BacktestEngine
from crypto_strategy_lab.support_resistance_analysis import generate_sr_analysis_reports


def cfg(**kw):
    base = dict(
        risk_mode=RiskMode.FIXED,
        fixed_r=10,
        initial_equity=1000,
        risk_per_leg=0.01,
        taker_fee=0,
        maker_fee=0,
        slippage=0,
        entry_mode=EntryMode.WAIT_UNTIL_CLOSED,
    )
    base.update(kw)
    return BacktestConfig(**base)


def _fake_engine(enable_sr):
    return SimpleNamespace(config=cfg(enable_support_resistance_analysis=enable_sr), _classify_sr_event_labels=BacktestEngine._classify_sr_event_labels)


def _fake_pos(near_support=False, near_resistance=False, support_held=False, resistance_held=False,
              support_state="APPROACHING_SUPPORT", resistance_state="APPROACHING_RESISTANCE"):
    return SimpleNamespace(
        sr_near_support=near_support, sr_near_resistance=near_resistance,
        sr_support_held=support_held, sr_resistance_held=resistance_held,
        sr_support_state=support_state, sr_resistance_state=resistance_state,
    )


def test_sr_event_labels_none_when_disabled():
    engine = _fake_engine(False)
    assert BacktestEngine._sr_event_labels(engine, _fake_pos()) is None


def test_sr_event_labels_no_nearby_sr_when_nothing_matches():
    engine = _fake_engine(True)
    labels = BacktestEngine._sr_event_labels(engine, _fake_pos())
    assert labels == ["NO_NEARBY_SR"]


def test_sr_event_labels_support_bounce_and_near_support():
    engine = _fake_engine(True)
    pos = _fake_pos(near_support=True, support_held=True)
    labels = BacktestEngine._sr_event_labels(engine, pos)
    assert "NEAR_SUPPORT" in labels
    assert "SUPPORT_BOUNCE" in labels
    assert "NO_NEARBY_SR" not in labels


def test_sr_event_labels_resistance_rejection():
    engine = _fake_engine(True)
    pos = _fake_pos(near_resistance=True, resistance_held=True)
    labels = BacktestEngine._sr_event_labels(engine, pos)
    assert labels == ["NEAR_RESISTANCE", "RESISTANCE_REJECTION"]


def test_sr_event_labels_resistance_breakout():
    engine = _fake_engine(True)
    pos = _fake_pos(resistance_state="RESISTANCE_BROKEN")
    labels = BacktestEngine._sr_event_labels(engine, pos)
    assert labels == ["RESISTANCE_BREAKOUT"]


def test_sr_event_labels_support_breakdown():
    engine = _fake_engine(True)
    pos = _fake_pos(support_state="SUPPORT_BROKEN")
    labels = BacktestEngine._sr_event_labels(engine, pos)
    assert labels == ["SUPPORT_BREAKDOWN"]


def _synthetic_trades():
    rows = []
    for i in range(6):
        rows.append({
            "side": "LONG",
            "long_sr_location": "NEAR_SUPPORT" if i % 2 == 0 else "BETWEEN_LEVELS",
            "long_sr_context": "SUPPORT_BOUNCE" if i % 2 == 0 else "NO_NEARBY_SR",
            "long_sr_support_distance_atr": 0.3 if i % 2 == 0 else 1.2,
            "long_sr_resistance_distance_atr": 2.0,
            "long_pair_net_r": 0.5 if i % 2 == 0 else -0.3,
            "long_pair_net_pnl": 5.0 if i % 2 == 0 else -3.0,
            "long_holding_minutes": 60,
            "market_regime": "BULL" if i < 3 else "SIDEWAYS",
            "short_sr_location": None,
        })
    return pd.DataFrame(rows)


def test_generate_sr_analysis_reports_creates_consolidated_workbook(tmp_path):
    trades = _synthetic_trades()
    reports = generate_sr_analysis_reports(trades, tmp_path)
    assert reports
    workbook = tmp_path / "support_resistance_analysis.xlsx"
    assert workbook.exists()
    assert load_workbook(workbook, read_only=True).sheetnames == ["Overview", "Location", "Distance", "Regime", "Event Context", "Hold Analysis", "Rejection Strength", "Test Count"]
    assert not (tmp_path / "sr_regime_analysis.csv").exists()
    assert not (tmp_path / "sr_distance_analysis.csv").exists()


def test_sr_event_context_report_has_expected_rows_and_columns(tmp_path):
    trades = _synthetic_trades()
    generate_sr_analysis_reports(trades, tmp_path)
    report = pd.read_excel(tmp_path / "support_resistance_analysis.xlsx", sheet_name="Event Context")
    assert set(report["context"]) == {"SUPPORT_BOUNCE", "NO_NEARBY_SR"}
    for column in ("trade_count", "winners", "losers", "win_rate", "total_r", "avg_r", "total_pnl", "avg_pnl"):
        assert column in report.columns
